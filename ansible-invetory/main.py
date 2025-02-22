import paramiko
import openpyxl
import socket
import re
import ipaddress
import subprocess
import os

# Define your sudo password in a file in the same directory called sudopass
SUDO_PASSWORD = "password"
timeout = "1s"
with open(f'{os.getcwd()}/sudopass') as f: 
    SUDO_PASSWORD = f.read()

with open('sudopass') as f: s = f.read()

def run_sudo_command(command):
    # Run a sudo command with password input
    process = subprocess.Popen(
        ['sudo', '-S'] + command,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True
    )
    stdout, stderr = process.communicate(input=SUDO_PASSWORD + '\n')
    return stdout, stderr

def get_ssh(ip):
    # Use system call to nmap to check if port 22 is open with a timeout
    stdout, stderr = run_sudo_command(['nmap', '-T', '4', '-p', '22', '--open', '--host-timeout', timeout, ip])
    return "22/tcp open" in stdout

def get_dns_record(ip):
    try:
        return socket.gethostbyaddr(ip)[0]
    except socket.herror:
        return "No DNS record"

def get_mac_address(ip):
    # Use system call to nmap to perform an ARP scan with a timeout
    stdout, stderr = run_sudo_command(['nmap', '-T', '4', '-sn', '--host-timeout', timeout, ip])
    
    # Parse the output to find the MAC address
    lines = stdout.splitlines()
    for line in lines:
        if "MAC Address:" in line:
            return line.split("MAC Address:")[1].strip()
    return None

def get_os_info(ip):
    # Use system call to nmap to perform OS detection with a timeout
    stdout, stderr = run_sudo_command(['nmap', '-T', '4', '-O', '-Pn', '--host-timeout', timeout, ip])
    
    # Parse the output to find OS information
    lines = stdout.splitlines()
    os_info = []
    for line in lines:
        if "OS details:" in line:
            os_info.append(line.split("OS details:")[1].strip())
    return os_info if os_info else None

def create_excel_report(data, filename):
    wb = openpyxl.Workbook()
    for ip_range, results in data.items():
        ws = wb.create_sheet(title=ip_range)
        ws.append(['IP', 'MAC Address', 'Hostname', 'OS','SSH'])
        for result in results:
            # Convert OS list to a string
            if isinstance(result[3], list):
                result[3] = ", ".join(result[3])  # Join list elements into a single string
            ws.append(result)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    wb.save(filename)



ip_ranges = [
    #"10.10.0.0/16",
    "10.10.20.0/24",
    "10.30.0.0/24",
    "10.40.0.0/24",
    "10.101.84.128/25",
    "10.101.84.0/25",
    "10.101.85.128/25",
    "10.101.86.0/23",
    "10.101.85.128/25",
    "10.101.148.0/22",
    "194.117.20.192/26"
    ]
all_results = {}


for ip_range in ip_ranges:
    ip_range_list = [str(ip) for ip in ipaddress.IPv4Network(ip_range.strip())]
    results = []
    for ip in ip_range_list:
        print(f"Checking {ip}")
        mac         = get_mac_address(ip)
        ssh_status  = get_ssh(ip)
        hostname    = get_dns_record(ip)
        os_info     = get_os_info(ip)
        
        results.append([ip, mac, hostname, os_info, ssh_status])
        print(f"Found {ip, mac, hostname, os_info, ssh_status}")
        all_results[ip_range] = results

    create_excel_report(all_results, f'range_{ip_range}.xlsx')
    print(f"Report generated as 'range_{ip_range}.xlsx'")

print("Scan Completed")